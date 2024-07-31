"""
data_preparation.py
-------------------
Contains functions for downloading and preparing data.
"""

# SPDX-FileCopyrightText: [2024] University of Manchester

# SPDX-License-Identifier: apache-2.0

import os
import requests
import json
import re
from openpyxl import load_workbook
from datasets import Dataset
from datasets import DatasetDict

def download_excel_file(url, local_path):
    """
    Download an Excel file from a given URL if it does not exist locally.

    Parameters:
    - url (str): URL of the Excel file.
    - local_path (str): Local path where the file will be saved.
    """
    if not os.path.exists(local_path):
        response = requests.get(url)
        if response.status_code == 200:
            with open(local_path, 'wb') as file:
                file.write(response.content)
            print(f"File downloaded successfully and saved as {local_path}")
        else:
            print(f"Failed to download file from {url}. Status code: {response.status_code}")
    else:
        print(f"File already exists at {local_path}")

def cleanText(text):
    """
    Clean text by removing non-ASCII characters.

    Parameters:
    - text (str): Text to be cleaned.

    Returns:
    - str: Cleaned text.
    """
    return re.sub(r'[^\x00-\x7F]+', '', text)

def generateDualPrompt(dataPoint):
    """
    Generate a dual prompt from a data point.

    Parameters:
    - dataPoint (dict): Dictionary containing original and rewritten text.

    Returns:
    - str: Generated dual prompt.
    """
    text = (
        f"Original: {dataPoint['original_text']}\n"
        f"Transformed: {dataPoint['rewritten_text']}"
    )
    return text

def initialiseData(filePath, outputJson):
    """
    Initialize data from an Excel file and save it to a JSON file.

    Parameters:
    - filePath (str): Path to the Excel file.
    - outputJson (str): Path to the output JSON file.
    """
    workbook = load_workbook(filename=filePath)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            'context': generateDualPrompt({
                'original_text': cleanText(row[0]),
                'rewritten_text': cleanText(row[3])
            }),
            'answer': cleanText(row[1]),
            'question': "What is the prompt that modifies original_text to rewritten_text"
        })

    for prompt in data[:5]:
        print(prompt)
        print("---------")

    with open(outputJson, 'w') as json_file:
        json.dump(data, json_file)
        print(f"Data from {filePath} saved to {outputJson}")

def tokenize_function(examples, tokenizer):
    """
    Tokenize the input examples using the provided tokenizer.

    Parameters:
    - examples (dict): Dictionary of examples to be tokenized.
    - tokenizer (AutoTokenizer): Tokenizer to be used.

    Returns:
    - dict: Tokenized examples.
    """
    return tokenizer(examples["question"], examples['context'])

def prepare_dataset(dataset, tokenizer):
    """
    Prepare the dataset by tokenizing it.

    Parameters:
    - dataset (DatasetDict): Dataset to be tokenized.
    - tokenizer (AutoTokenizer): Tokenizer to be used.

    Returns:
    - DatasetDict: Tokenized dataset.
    """
    tokenized_dataset = dataset.map(lambda examples: tokenize_function(examples, tokenizer), batched=True)
    return tokenized_dataset

def load_my_dataset(json_file):
    """
    Load a dataset from a JSON file.

    Parameters:
    - json_file (str): Path to the JSON file.

    Returns:
    - Dataset: Loaded dataset.
    """
    with open(json_file, 'r') as json_file:
        data = json.load(json_file)

    return Dataset.from_dict({
        "question": [x['question'] for x in data],
        "context": [x['context'] for x in data],
        "answer": [x['answer'] for x in data]
    })

def my_dataset(train_json='train.json', test_json='test.json'):
    """
    Load the train and test datasets.

    Parameters:
    - train_json (str): Path to the train JSON file.
    - test_json (str): Path to the test JSON file.

    Returns:
    - DatasetDict: Dictionary containing the train and test datasets.
    """
    train = load_my_dataset(train_json)
    test = load_my_dataset(test_json)
    return DatasetDict({"train": train, "test": test})

def formatting_func(input):
    """
    Format the input for the model.

    Parameters:
    - input (dict): Input dictionary.

    Returns:
    - list: List containing the formatted text.
    """
    text = f"Question: {input['question'][0]},\nContext: {input['context'][0]}\nAnswer: {input['answer'][0]}"
    return [text]